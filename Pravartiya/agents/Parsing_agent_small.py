#!/usr/bin/env python
# agents/newsletter_parsing_agent.py
# ============================================================
# PRAVARTIYA NEWSLETTER PARSING AGENT
# ============================================================

import sys
from pathlib import Path

BASE_DIR = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(BASE_DIR))

from storage.minio_client import MinIOClient

import os
import re
import json
import logging
import warnings
import time
from datetime import datetime
from pathlib import Path

import pandas as pd
import torch

from dotenv import load_dotenv
from openpyxl import load_workbook, Workbook
from unstructured.partition.pdf import partition_pdf

from langchain_community.llms import Ollama
from langchain.prompts import PromptTemplate

# ============================================================
# LOGGING
# ============================================================

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    handlers=[logging.StreamHandler()]
)

warnings.filterwarnings("ignore")
load_dotenv()

# ============================================================
# GPU
# ============================================================

device = torch.device("cuda" if torch.cuda.is_available() else "cpu")

if device.type == "cuda":

    os.environ["OLLAMA_USE_GPU"] = "1"
    os.environ["OLLAMA_NUM_GPU_LAYERS"] = "35"

    print(f"Using GPU: {torch.cuda.get_device_name(0)}")

else:

    os.environ["OLLAMA_USE_GPU"] = "0"
    print("Using CPU")

# ============================================================
# LLM
# ============================================================

llm = Ollama(model="mistral:latest")

# ============================================================
# PATHS
# ============================================================

DATA_DIR = BASE_DIR / "data"

OUTPUT_EXCEL_DIR = DATA_DIR / "output_excels"
OUTPUT_EXCEL_DIR.mkdir(parents=True, exist_ok=True)

CREATED_EXCELS = set()

# ============================================================
# MONTH FOLDER
# ============================================================

def get_month_folder() -> Path:

    month_json = DATA_DIR / "month_range.json"

    if not month_json.exists():

        raise RuntimeError(
            "month_range.json not found"
        )

    with open(month_json, "r") as f:

        month = json.load(f)

    ms = datetime.strptime(
        month["month_start"],
        "%Y-%m-%d"
    )

    me = datetime.strptime(
        month["month_end"],
        "%Y-%m-%d"
    )

    folder = OUTPUT_EXCEL_DIR / (
        f"newsletter_{ms:%Y-%m-%d}_to_{me:%Y-%m-%d}"
    )

    if folder.exists():

        import shutil
        shutil.rmtree(folder)

    folder.mkdir(parents=True)

    return folder


MONTH_FOLDER = get_month_folder()

logging.info(
    f"Monthly newsletter folder -> {MONTH_FOLDER}"
)

# ============================================================
# TITLE FILTER
# ============================================================

AMENDED_TITLE_PATTERN = re.compile(
    r'last\s+amended\s+on|amended\s+as\s+on',
    re.IGNORECASE
)

def is_amended_title(title: str) -> bool:

    if not isinstance(title, str):

        return False

    return bool(
        AMENDED_TITLE_PATTERN.search(title)
    )

# ============================================================
# EFFECTIVE DATE DETECTION
# ============================================================

GAZETTE_FORCE_PATTERN = re.compile(
    r'they\s+shall\s+come\s+into\s+force\s+on\s+the\s+date\s+of\s+their\s+publication\s+in\s+the\s+official\s+gazette',
    re.IGNORECASE
)

_CITY_PAT = (
    r"(?:mumbai|bombay|new\s+delhi|delhi|hyderabad|"
    r"chennai|madras|kolkata|calcutta|bangalore|"
    r"bengaluru|ahmedabad|pune)"
)

NOTIFICATION_DATE_PATTERNS = [

    re.compile(
        _CITY_PAT + r",?\s+the\s+(\d{1,2}(?:st|nd|rd|th)?\s+[A-Za-z]+,?\s*\d{4})",
        re.IGNORECASE
    ),

    re.compile(
        r"\b(\d{1,2}(?:st|nd|rd|th)?\s+[A-Za-z]+\s*,?\s*\d{4})\b",
        re.IGNORECASE
    ),
]

def extract_notification_date(text: str) -> str:

    sample = text[:3000]

    for pattern in NOTIFICATION_DATE_PATTERNS:

        match = pattern.search(sample)

        if match:

            return match.group(1).strip()

    return "N/A"


def gazette_force_present(text: str) -> bool:

    return bool(
        GAZETTE_FORCE_PATTERN.search(text)
    )


def determine_effective_date(text: str) -> str:

    if gazette_force_present(text):

        return extract_notification_date(text)

    return "N/A"

# ============================================================
# PDF EXTRACTION
# ============================================================

def extract_pdf_text(pdf_path: Path) -> str:

    raw = partition_pdf(
        filename=str(pdf_path),
        strategy="fast",
        include_page_breaks=False
    )

    text = "\n".join(
        str(el) for el in raw if el
    ).strip()

    if not text:

        logging.info("Fallback to hi_res OCR")

        raw = partition_pdf(
            filename=str(pdf_path),
            strategy="hi_res"
        )

        text = "\n".join(
            str(el) for el in raw if el
        ).strip()

    upper_text = text.upper()

    english_start = upper_text.find(
        "SECURITIES AND EXCHANGE BOARD OF INDIA"
    )

    if english_start != -1:

        text = text[english_start:]

    return text

# ============================================================
# REGULATION CORE EXTRACTION
# ============================================================

def extract_regulation_core(text: str) -> str:

    lines = text.splitlines()

    keep = []

    capture = False

    for line in lines:

        clean = line.strip()

        if not clean:
            continue

        if re.search(
            r'in exercise of the powers conferred',
            clean,
            re.IGNORECASE
        ):

            capture = True
            continue

        if not capture:
            continue

        if re.search(
            r'(first|second|third|fourth|fifth|sixth|seventh|eighth)\s+amendment',
            clean,
            re.IGNORECASE
        ):
            continue

        if re.search(
            r'as amended upto|as amended up to',
            clean,
            re.IGNORECASE
        ):
            continue

        if re.match(
            r'^\([a-z]+\)',
            clean
        ):
            continue

        if len(clean) > 10:

            keep.append(clean)

        if len(keep) >= 4000:

            break

    return "\n".join(keep)

# ============================================================
# CLEANER
# ============================================================

def clean_summary(summary: str) -> str:

    if not summary:

        return "NA"

    summary = re.sub(
        r'https?://\S+',
        '',
        summary
    )

    summary = re.sub(
        r'\S+@\S+',
        '',
        summary
    )

    summary = re.sub(
        r'circular\s+no\.?\s*[:\-]?\s*\S+',
        '',
        summary,
        flags=re.IGNORECASE
    )

    summary = re.sub(
        r'\n\s*\n+',
        '\n',
        summary
    )

    return summary.strip()

# ============================================================
# NEWSLETTER PROMPT
# ============================================================

# NEWSLETTER_REGULATION_PROMPT = PromptTemplate(
#     template="""
# You are a senior SEBI regulatory analyst preparing a Pravarthiya newsletter summary.

# The document is a SEBI Regulation amendment/update.

# The summary is meant for:
# - listed entities
# - compliance officers
# - intermediaries
# - legal teams

# MANDATORY REQUIREMENTS:

# 1. The summary MUST include:
# - Date of circular
# - Effective date
# - Gist of amendment
# - Existing provisions prior to amendment
# - Practical compliance implication

# 2. Start the summary with wording such as:
# - "SEBI has issued this circular and introduced..."
# - "SEBI has issued this circular and amended..."
# - "SEBI has issued this circular and changed..."

# 3. Clearly explain:
# - what changed
# - what existed earlier
# - what is the new requirement
# - who is impacted

# 4. If there is an action point for listed entities/intermediaries,
# mention it at the end.

# 5. STRICTLY AVOID:
# - circular index references
# - clause numbering
# - vague legal wording
# - email IDs
# - procedural boilerplate
# - copying regulation text verbatim

# 6. Use concise newsletter-style language.

# 7. Output ONLY bullet points.

# 8. Maximum 6 bullet points.

# ISSUE DATE:
# {issue_date}

# EFFECTIVE DATE:
# {effective_date}

# TEXT:
# {text}

# FINAL SUMMARY:
# """,
#     input_variables=[
#         "issue_date",
#         "effective_date",
#         "text"
#     ]
# )

NEWSLETTER_REGULATION_PROMPT = PromptTemplate(
    template="""
You are a senior SEBI regulatory analyst preparing a Pravarthiya newsletter summary.

The document is a SEBI Regulation amendment/update.

The summary is meant for:
- listed entities
- compliance officers
- intermediaries
- legal teams

MANDATORY REQUIREMENTS:

1. Summarise ONLY the material regulatory amendments.

2. Start the summary with wording such as:
- "SEBI has issued this circular and amended..."
- "SEBI has issued this circular and introduced..."
- "SEBI has issued this circular and changed..."

3. Focus ONLY on:
- material compliance changes
- governance changes
- operational changes
- disclosure changes
- threshold changes
- investor impact
- listed entity obligations

4. Clearly explain:
- what existed earlier
- what changed now
- practical compliance impact

5. If there is an action point for listed entities/intermediaries,
mention it at the end.

6. STRICTLY AVOID:
- circular index references
- clause numbering
- drafting corrections
- punctuation changes
- wording substitutions
- grammar edits
- capitalization edits
- vague legal wording
- email IDs
- procedural boilerplate
- markdown headings
- title generation
- effective date generation
- repeating issue date
- copying regulation text verbatim

7. Ignore editorial or drafting amendments unless they materially change compliance obligations.

8. Use concise newsletter-style language.

9. Output ONLY bullet points.

10. Maximum 6 bullet points.

11. Prioritize ONLY amendments that materially impact:
- compliance process
- governance obligations
- disclosure requirements
- investor servicing
- monetary thresholds
- operational workflow

Do not describe every amendment individually.
Combine related amendments into a single business-level summary point.
TEXT:
{text}

FINAL SUMMARY:
""",
    input_variables=["text"]
)

# ============================================================
# SUMMARY GENERATION
# ============================================================

# def generate_newsletter_regulation_summary(
#     text: str,
#     issue_date: str,
#     effective_date: str
# ):

#     core_text = extract_regulation_core(text)

#     core_text = core_text[:12000]

#     summary = llm.invoke(
#         NEWSLETTER_REGULATION_PROMPT.format(
#             issue_date=issue_date,
#             effective_date=effective_date,
#             text=core_text
#         )
#     ).strip()

#     return clean_summary(summary)


def generate_newsletter_regulation_summary(
    text: str
):

    core_text = extract_regulation_core(text)

    core_text = core_text[:12000]

    summary = llm.invoke(
        NEWSLETTER_REGULATION_PROMPT.format(
            text=core_text
        )
    ).strip()

    return clean_summary(summary)

# ============================================================
# EXCEL UPDATE
# ============================================================

def update_excel(row: pd.Series):

    vertical = row["Verticals"]

    sub = row["SubCategory"]

    excel_path = (
        MONTH_FOLDER /
        f"{vertical}_Newsletter.xlsx"
    )

    CREATED_EXCELS.add(excel_path.name)

    if excel_path.exists():

        wb = load_workbook(excel_path)

    else:

        wb = Workbook()

        wb.remove(wb.active)

    sheet_name = sub if sub else "Regulations"

    if sheet_name not in wb.sheetnames:

        ws = wb.create_sheet(title=sheet_name)

        ws.append(list(row.index))

    else:

        ws = wb[sheet_name]

    ws.append([
        row.get(c, "NA")
        for c in row.index
    ])

    wb.save(excel_path)

    wb.close()

# ============================================================
# PROCESS ROW
# ============================================================

def process_newsletter_row(
    row: pd.Series
) -> pd.Series | None:

    sub = row.get("SubCategory", "")

    if not isinstance(sub, str):

        return None

    if sub.strip().lower() != "regulations":

        return None

    title = row.get("Title", "")

    if is_amended_title(title):

        logging.info(
            f"Skipping amended-title PDF: {title}"
        )

        return None

    pdf_path = Path(row["Path"])

    try:

        text = extract_pdf_text(pdf_path)

    except Exception as e:

        logging.error(
            f"PDF extraction failed: {e}"
        )

        row["Summary"] = "NA"

        return row

    effective_date = determine_effective_date(text)

    issue_date = ""

    for col in [
        "Date",
        "IssueDate",
        "Published",
        "PublishedDate",
        "Issue Date"
    ]:

        val = row.get(col, "")

        if val and str(val).strip():

            if isinstance(val, datetime):

                issue_date = val.strftime(
                    "%b %d, %Y"
                )

            else:

                issue_date = str(val).strip()

            break

    # summary = generate_newsletter_regulation_summary(
    #     text=text,
    #     issue_date=issue_date,
    #     effective_date=effective_date
    # )

    summary = generate_newsletter_regulation_summary(
        text=text
    )


    # row["Summary"] = summary
    final_summary = (
        f"{title} dated {issue_date}\n"
        f"Effective date - {effective_date}\n\n"
        f"{summary}"
    )

    row["Summary"] = final_summary
    row["EffectiveDate"] = effective_date

    row["EmbeddingText"] = text[:8000]

    return row

# ============================================================
# MAIN
# ============================================================

def main(excel_file: str):

    df = pd.read_excel(excel_file)

    required = [
        "Verticals",
        "SubCategory",
        "Path"
    ]

    for col in required:

        if col not in df.columns:

            raise ValueError(
                f"Missing required column: {col}"
            )

    if "EffectiveDate" not in df.columns:

        df["EffectiveDate"] = ""

    logging.info(
        f"Total rows in input: {len(df)}"
    )

    sebi_mask = df["Verticals"].str.strip().str.lower().isin(
        {
            "listed companies",
            "sebi",
            "aif"
        }
    )

    df_sebi = df[sebi_mask].copy()

    logging.info(
        f"SEBI rows to process: {len(df_sebi)}"
    )

    start = time.time()

    processed_count = 0

    for idx, row in df_sebi.iterrows():

        logging.info(
            f"[{idx+1}] Processing: "
            f"{row.get('Title', '')[:80]}"
        )

        processed = process_newsletter_row(row)

        if processed is None:

            continue

        update_excel(processed)

        processed_count += 1

    logging.info(
        f"Completed processing "
        f"{processed_count} regulations "
        f"in {time.time() - start:.2f}s"
    )

    # ========================================================
    # MINIO UPLOAD
    # ========================================================

    try:

        minio = MinIOClient()

        month_folder_name = MONTH_FOLDER.name

        minio_prefix = (
            f"monthly_outputs/{month_folder_name}/"
        )

        minio.delete_prefix(minio_prefix)

        for excel_name in CREATED_EXCELS:

            local_excel = MONTH_FOLDER / excel_name

            object_path = (
                f"{minio_prefix}{excel_name}"
            )

            minio.upload_file(
                local_path=str(local_excel),
                object_path=object_path
            )

        logging.info(
            f"Uploaded newsletter files to MinIO"
        )

    except Exception as e:

        logging.error(
            f"MinIO upload failed: {e}"
        )

# ============================================================
# ENTRY
# ============================================================

if __name__ == "__main__":

    excel = DATA_DIR / "Searching_agent_output.xlsx"

    if not excel.exists():

        raise FileNotFoundError(
            "Searching_agent_output.xlsx not found"
        )

    main(str(excel))

