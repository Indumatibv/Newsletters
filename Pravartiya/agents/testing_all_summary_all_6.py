import json
import requests
import re
import openpyxl

from openpyxl.styles import (
    Font,
    PatternFill,
    Alignment,
    Border,
    Side
)

from openpyxl.utils import (
    get_column_letter
)


# ============================================================
# STRICT INLINE FOOTNOTE CLAUSE EXTRACTION
# ============================================================

def extract_clause_with_footnote_marker(
    text,
    footnote_num
):

    marker = f"{footnote_num}["

    if marker not in text:
        return text

    # ========================================================
    # Extract ONLY exact sentence containing footnote marker
    # Prevents neighboring amendment contamination
    # ========================================================

    pattern = rf'([^.]*?{re.escape(marker)}.*?[.])'

    match = re.search(
        pattern,
        text,
        flags=re.DOTALL
    )

    if match:
        return match.group(1).strip()

    return text


# ============================================================
# DETECT AMENDMENT TYPE
# ============================================================

def detect_amendment_type(
    footer_text
):

    footer_lower = footer_text.lower()

    if "inserted" in footer_lower:
        return "inserted"

    elif "omitted" in footer_lower:
        return "omitted"

    elif "substituted" in footer_lower:
        return "substituted"

    return "unknown"


# ============================================================
# MAIN PROCESSOR
# ============================================================

def process_all_footers_to_excel(

    mapped_json_path,

    output_excel_path

):

    print(
        f"Loading dataset from '{mapped_json_path}'..."
    )

    with open(
        mapped_json_path,
        "r",
        encoding="utf-8"
    ) as f:

        mapped_data = json.load(f)

    if not mapped_data:

        print("Error: Dataset is empty.")

        return


    # ========================================================
    # EXCEL SETUP
    # ========================================================

    wb = openpyxl.Workbook()

    ws = wb.active

    ws.title = "Regulatory Summary Matrix"

    ws.views.sheetView[0].showGridLines = True


    # ========================================================
    # STYLING
    # ========================================================

    navy_header_fill = PatternFill(
        start_color="1F497D",
        end_color="1F497D",
        fill_type="solid"
    )

    white_bold_font = Font(
        name="Segoe UI",
        size=11,
        bold=True,
        color="FFFFFF"
    )

    regular_font = Font(
        name="Segoe UI",
        size=10
    )

    thin_border_side = Side(
        border_style="thin",
        color="D9D9D9"
    )

    thin_border = Border(
        left=thin_border_side,
        right=thin_border_side,
        top=thin_border_side,
        bottom=thin_border_side
    )

    zebra_fill = PatternFill(
        start_color="F9FAFB",
        end_color="F9FAFB",
        fill_type="solid"
    )


    headers = [

        "Footer ID",

        "Amendment Type",

        "Footer Text",

        "Matched Boundaries",

        "Mapped Regulation Chunks Content",

        "Filtered Amendment Context",

        "Mistral Compliance Summary"
    ]


    ws.append(headers)


    for col_idx, header in enumerate(headers, 1):

        cell = ws.cell(
            row=1,
            column=col_idx
        )

        cell.fill = navy_header_fill

        cell.font = white_bold_font

        cell.alignment = Alignment(
            horizontal="center",
            vertical="center",
            wrap_text=True
        )

    ws.row_dimensions[1].height = 28


    # ========================================================
    # PROCESS FOOTERS
    # ========================================================

    row_num = 2

    for footer_id, payload in mapped_data.items():

        footer_text = payload.get(
            "footer_text",
            ""
        ).strip()


        amendment_type = detect_amendment_type(
            footer_text
        )


        boundaries_list = payload.get(
            "matched_boundaries",
            payload.get(
                "matched_scopes",
                []
            )
        )

        boundaries_str = "\n".join(
            boundaries_list
        )


        regulation_chunks = payload.get(
            "mapped_regulation_chunks",
            []
        )


        full_chunks_text_list = []

        extracted_clauses_list = []


        # ====================================================
        # EXTRACT LOCALIZED CLAUSES
        # ====================================================

        for idx, chunk in enumerate(
            regulation_chunks,
            1
        ):

            chunk_text = chunk.get(
                "text",
                ""
            ).strip()

            chapter = chunk.get(
                "chapter",
                "N/A"
            )

            section = chunk.get(
                "section",
                "N/A"
            )

            ch_info = (
                f"Chunk {idx} "
                f"({chapter} -> Section {section}):\n"
                f"{chunk_text}"
            )

            full_chunks_text_list.append(
                ch_info
            )


            isolated = (
                extract_clause_with_footnote_marker(
                    chunk_text,
                    footer_id
                )
            )

            extracted_clauses_list.append(
                isolated
            )


        chunks_column_data = (
            "\n\n---\n\n".join(
                full_chunks_text_list
            )
        )


        # ====================================================
        # PRESERVE ORDER + REMOVE DUPLICATES
        # ====================================================

        filtered_context = "\n\n".join(
            dict.fromkeys(
                extracted_clauses_list
            )
        )


        # ====================================================
        # STRICT LEGAL PROMPT
        # ====================================================

        prompt = f"""
[SYSTEM INSTRUCTION]

You are a factual legal documentation parser.

You must generate a compliance summary strictly from the provided contexts.

Do not infer.
Do not speculate.
Do not introduce external legal interpretation.
Do not generate broader governance implications unless explicitly mentioned.

=======================================================
AMENDMENT TYPE
=======================================================

{amendment_type}

=======================================================
CONTEXT A: TARGET FOOTNOTE
=======================================================

{footer_text}

=======================================================
CONTEXT B: TARGET REGULATION CLAUSE
=======================================================

{filtered_context}

=======================================================
INTERPRETATION RULES
=======================================================

- CONTEXT A contains amendment metadata and, where available, the prior legal provision before amendment.

- CONTEXT B contains the currently applicable amended regulation text.

- Use both contexts together to identify the practical legal effect of the amendment.

- For substituted provisions, compare the prior provision in CONTEXT A with the amended provision in CONTEXT B and summarize only the material regulatory change.

- For omitted provisions, identify which requirement, obligation, exemption, disclosure, or timeline stands removed.

- For inserted provisions, identify the newly introduced requirement, obligation, disclosure, exemption, timeline, or compliance condition.

- Do not mechanically compare old and new wording.

- Focus only on the material legal and compliance impact reflected in the text.

=======================================================
MANDATORY OUTPUT FORMAT
=======================================================

Your output must contain ONLY the following three fields.

Do not include:
- introductory lines
- dates
- circular references
- section numbers
- administrative references
- email ids
- commentary
- assumptions
- meta-analysis

-------------------------------------------------------

Gist of amendment:
(State only the exact legal and practical effect of the amendment reflected in the text.

- If a provision is inserted, clearly state the new requirement, obligation, disclosure, exemption, timeline, or compliance condition introduced.

- If a provision is omitted, clearly state that the relevant requirement, obligation, exemption, disclosure requirement, or timeline stands removed.

- If a provision is substituted, summarize the practical regulatory change introduced by the revised provision instead of mechanically comparing old and new wording.

- Focus only on the material change introduced by the amendment.

- Avoid generic statements such as "specified by the Board" where such wording already existed in the earlier provision.

Do not speculate beyond the text.)

Existing provisions of Law prior to amendment:
(Extract only the prior legal provision from the footer text if available.
If not available, write:
"Not explicitly mentioned")

Action point for listed entity if any:
(State only compliance actions directly resulting from the amendment text.

If a provision is omitted, state that entities should update internal governance records, trackers, timelines, exemptions, or compliance references to remove references to the omitted provision.

Do not speculate beyond the amendment text.)
"""


        # ====================================================
        # OLLAMA REQUEST
        # ====================================================

        OLLAMA_URL = (
            "http://localhost:11434/api/generate"
        )

        api_payload = {

            "model": "mistral:latest",

            "prompt": prompt,

            "stream": False,

            "options": {

                "temperature": 0.0,

                "top_p": 0.1
            }
        }


        print(
            f"Generating summary for Footer ID {footer_id}..."
        )


        summary_result = ""


        try:

            response = requests.post(
                OLLAMA_URL,
                json=api_payload,
                timeout=90
            )

            if response.status_code == 200:

                summary_result = response.json().get(
                    "response",
                    ""
                ).strip()

            else:

                summary_result = (
                    f"[ERROR] Status code: "
                    f"{response.status_code}"
                )

        except Exception as e:

            summary_result = (
                f"[ERROR] Ollama failure: {str(e)}"
            )


        # ====================================================
        # WRITE EXCEL ROW
        # ====================================================

        row_data = [

            footer_id,

            amendment_type,

            footer_text,

            boundaries_str,

            chunks_column_data,

            filtered_context,

            summary_result
        ]


        ws.append(row_data)


        current_fill = (
            zebra_fill
            if row_num % 2 == 0
            else PatternFill(fill_type=None)
        )


        for col_idx in range(1, 8):

            cell = ws.cell(
                row=row_num,
                column=col_idx
            )

            cell.font = regular_font

            cell.border = thin_border

            cell.fill = current_fill

            cell.alignment = Alignment(
                horizontal="left",
                vertical="top",
                wrap_text=True
            )


        row_num += 1


    # ========================================================
    # COLUMN WIDTHS
    # ========================================================

    column_widths = {

        1: 12,

        2: 16,

        3: 40,

        4: 20,

        5: 55,

        6: 45,

        7: 60
    }


    for col_idx, width in column_widths.items():

        ws.column_dimensions[
            get_column_letter(col_idx)
        ].width = width


    wb.save(output_excel_path)


    print(
        f"\nSpreadsheet generated successfully:\n"
        f"{output_excel_path}"
    )


# ============================================================
# RUN
# ============================================================

if __name__ == "__main__":

    process_all_footers_to_excel(

        mapped_json_path="footers_mapped_to_regulations.json",

        output_excel_path="SEBI_Compliance_Summary_Matrix.xlsx"
    )
