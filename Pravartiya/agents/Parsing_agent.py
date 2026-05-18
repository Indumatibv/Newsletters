#!/usr/bin/env python
# agents/newsletter_parsing_agent.py
# ============================================================
# PRAVARTIYA NEWSLETTER PARSING AGENT
# ============================================================

import sys
from pathlib import Path

BASE_DIR = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(BASE_DIR))

from storage.minio_client import MinIOClient

import os
import re
import json
import logging
import warnings
import time
from datetime import datetime
from pathlib import Path

import pandas as pd
import torch

from dotenv import load_dotenv
from openpyxl import load_workbook, Workbook
from unstructured.partition.pdf import partition_pdf

from langchain_community.llms import Ollama
from langchain.prompts import PromptTemplate

# ============================================================
# LOGGING
# ============================================================

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    handlers=[logging.StreamHandler()]
)

warnings.filterwarnings("ignore")

load_dotenv()

# ============================================================
# GPU
# ============================================================

device = torch.device(
    "cuda" if torch.cuda.is_available() else "cpu"
)

if device.type == "cuda":

    os.environ["OLLAMA_USE_GPU"] = "1"
    os.environ["OLLAMA_NUM_GPU_LAYERS"] = "35"

    print(
        f"Using GPU: "
        f"{torch.cuda.get_device_name(0)}"
    )

else:

    os.environ["OLLAMA_USE_GPU"] = "0"

    print("Using CPU")

# ============================================================
# LLM
# ============================================================

llm = Ollama(model="mistral:latest")

# ============================================================
# PATHS
# ============================================================

DATA_DIR = BASE_DIR / "data"

OUTPUT_EXCEL_DIR = DATA_DIR / "output_excels"

OUTPUT_EXCEL_DIR.mkdir(
    parents=True,
    exist_ok=True
)

CREATED_EXCELS = set()

# ============================================================
# MONTH FOLDER
# ============================================================

def get_month_folder() -> Path:

    month_json = DATA_DIR / "month_range.json"

    if not month_json.exists():

        raise RuntimeError(
            "month_range.json not found"
        )

    with open(month_json, "r") as f:

        month = json.load(f)

    ms = datetime.strptime(
        month["month_start"],
        "%Y-%m-%d"
    )

    me = datetime.strptime(
        month["month_end"],
        "%Y-%m-%d"
    )

    folder = OUTPUT_EXCEL_DIR / (
        f"newsletter_{ms:%Y-%m-%d}"
        f"_to_{me:%Y-%m-%d}"
    )

    if folder.exists():

        import shutil

        shutil.rmtree(folder)

    folder.mkdir(parents=True)

    return folder


MONTH_FOLDER = get_month_folder()

logging.info(
    f"Monthly newsletter folder -> "
    f"{MONTH_FOLDER}"
)

# ============================================================
# PDF EXTRACTION
# ============================================================

def extract_pdf_text(
    pdf_path: Path
) -> str:

    raw = partition_pdf(
        filename=str(pdf_path),
        strategy="fast",
        include_page_breaks=False
    )

    text = "\n".join(
        str(el)
        for el in raw
        if el
    ).strip()

    if not text:

        logging.info(
            "Fallback to hi_res OCR"
        )

        raw = partition_pdf(
            filename=str(pdf_path),
            strategy="hi_res"
        )

        text = "\n".join(
            str(el)
            for el in raw
            if el
        ).strip()

    return text

# ============================================================
# CLEANER
# ============================================================

def clean_summary(
    summary: str
) -> str:

    if not summary:

        return "NA"

    summary = re.sub(
        r'https?://\S+',
        '',
        summary
    )

    summary = re.sub(
        r'\S+@\S+',
        '',
        summary
    )

    summary = re.sub(
        r'circular\s+no\.?\s*[:\-]?\s*\S+',
        '',
        summary,
        flags=re.IGNORECASE
    )

    summary = re.sub(
        r'(SEBI|CIR|CFD|HO)'
        r'/[A-Z0-9/_\-.]+',
        '',
        summary,
        flags=re.IGNORECASE
    )

    summary = re.sub(
        r'\n\s*\n+',
        '\n',
        summary
    )

    summary = re.sub(
        r'\(?www\.[^\s)]+\)?',
        '',
        summary,
        flags=re.IGNORECASE
    )
    return summary.strip()

# ============================================================
# PROMPTS
# ============================================================

INFORMAL_GUIDANCE_PROMPT = PromptTemplate(
    template="""
You are a senior SEBI regulatory analyst preparing a Pravarthiya newsletter summary.

The document is a SEBI Informal Guidance letter.

MANDATORY REQUIREMENTS:

1. The summary MUST clearly include:
- Background and relevant facts
- Query raised before SEBI
- SEBI's response/interpretation

2. Facts should focus ONLY on:
- regulatory issue
- legal interpretation
- compliance concern
- relevant SEBI regulation

3. Facts should NOT focus on:
- transaction mechanics
- percentage shareholding
- acquisition size
- commercial deal details
- numerical transaction details

4. Clearly explain:
- what clarification was sought
- what SEBI interpreted
- compliance implication

5. Use concise newsletter-style language.

6. Output ONLY bullet points.

7. Maximum 5 bullet points.

TEXT:
{text}

FINAL SUMMARY:
""",
    input_variables=["text"]
)


# EXCHANGE_CIRCULAR_PROMPT = PromptTemplate(
#     template="""
# You are a senior regulatory analyst preparing a Pravarthiya newsletter summary.

# The document is an NSE/BSE circular.

# MANDATORY REQUIREMENTS:

# 1. The summary MUST clearly include:
# - gist of circular
# - stated compliance implication
# - stated operational implication

# 2. Summarize ONLY information explicitly available in the circular text.

# 3. Do NOT infer, assume, or invent amendment details that are not clearly stated in the circular.

# 4. If the circular merely forwards or references another SEBI circular:
# - summarize the forwarding communication
# - summarize the stated purpose
# - summarize the stated compliance action only

# 5. Clearly explain:
# - what the circular is about
# - who is impacted
# - what entities/intermediaries are expected to do

# 6. STRICTLY AVOID:
# - hallucinated amendments
# - inferred regulatory changes
# - unsupported compliance obligations
# - circular reference numbers
# - procedural boilerplate
# - URLs
# - email IDs
# - repetitive legal wording
# - clause dumping
# - copying circular language verbatim

# 7. Use concise newsletter-style language.

# 8. Output ONLY bullet points.

# 9. Maximum 5 bullet points.

# 10. Focus ONLY on material information explicitly stated in the circular.

# TEXT:
# {text}

# FINAL SUMMARY:
# """,
#     input_variables=["text"]
# )

EXCHANGE_CIRCULAR_PROMPT = PromptTemplate(
    template="""
You are a senior regulatory analyst preparing a Pravarthiya newsletter summary.

The document is an NSE/BSE circular.

MANDATORY REQUIREMENTS:

1. The summary MUST include:
- gist of circular
- stated compliance action
- impacted entities

2. Summarize ONLY information explicitly available in the circular text.

3. Do NOT infer, assume, or invent:
- amendment details
- operational changes
- compliance obligations
- timelines
- thresholds
unless they are explicitly stated in the circular.

4. If the circular merely forwards or references another SEBI circular:
- clearly state that the exchange has referred/circulated the SEBI circular
- summarize only the stated purpose/reference
- mention necessary compliance action if stated

5. STRICTLY AVOID:
- circular reference numbers
- URLs
- email IDs
- procedural boilerplate
- unsupported assumptions
- hallucinated regulatory changes
- copying circular text verbatim
- mentioning absence of information
- speculating about referenced circulars


6. Use concise newsletter-style language.

7. Output ONLY bullet points.

8. Maximum 4 bullet points.

9. Keep the summary factual, concise and grounded strictly in the circular text.

TEXT:
{text}

FINAL SUMMARY:
""",
    input_variables=["text"]
)

# ============================================================
# SUMMARY GENERATORS
# ============================================================

def generate_informal_guidance_summary(
    text: str
):

    summary = llm.invoke(
        INFORMAL_GUIDANCE_PROMPT.format(
            text=text[:12000]
        )
    ).strip()

    return clean_summary(summary)


def generate_exchange_circular_summary(
    text: str
):

    summary = llm.invoke(
        EXCHANGE_CIRCULAR_PROMPT.format(
            text=text[:12000]
        )
    ).strip()

    return clean_summary(summary)

# ============================================================
# EXCEL UPDATE
# ============================================================

def update_excel(
    row: pd.Series
):

    vertical = row["Verticals"]

    sub = row["SubCategory"]

    excel_path = (
        MONTH_FOLDER /
        f"{vertical}_Newsletter.xlsx"
    )

    CREATED_EXCELS.add(
        excel_path.name
    )

    if excel_path.exists():

        wb = load_workbook(
            excel_path
        )

    else:

        wb = Workbook()

        wb.remove(
            wb.active
        )

    sheet_name = (
        sub
        if sub
        else "General"
    )

    if sheet_name not in wb.sheetnames:

        ws = wb.create_sheet(
            title=sheet_name
        )

        ws.append(
            list(row.index)
        )

    else:

        ws = wb[sheet_name]

    ws.append([
        row.get(c, "NA")
        for c in row.index
    ])

    wb.save(excel_path)

    wb.close()

# ============================================================
# ISSUE DATE
# ============================================================

def get_issue_date(
    row: pd.Series
) -> str:

    issue_date = ""

    for col in [
        "Date",
        "IssueDate",
        "Published",
        "PublishedDate",
        "Issue Date"
    ]:

        val = row.get(col, "")

        if val and str(val).strip():

            if isinstance(
                val,
                datetime
            ):

                issue_date = val.strftime(
                    "%b %d, %Y"
                )

            else:

                issue_date = str(val).strip()

            break

    return issue_date

# ============================================================
# PROCESSOR
# ============================================================

def process_newsletter_row(
    row: pd.Series
) -> pd.Series | None:

    sub = row.get(
        "SubCategory",
        ""
    )

    if not isinstance(
        sub,
        str
    ):

        return None

    sub_clean = (
        sub.strip().lower()
    )

    allowed = {
        "informal guidance",
        "circular-bse",
        "circular-nse"
    }

    if sub_clean not in allowed:

        return None

    pdf_path = Path(
        row["Path"]
    )

    try:

        text = extract_pdf_text(
            pdf_path
        )

    except Exception as e:

        logging.error(
            f"PDF extraction failed: {e}"
        )

        row["Summary"] = "NA"

        return row

    title = row.get(
        "Title",
        ""
    )

    issue_date = get_issue_date(
        row
    )

    # ========================================================
    # INFORMAL GUIDANCE
    # ========================================================

    if sub_clean == "informal guidance":

        summary = (
            generate_informal_guidance_summary(
                text=text
            )
        )

        final_summary = (
            f"{title} dated "
            f"{issue_date}\n\n"
            f"{summary}"
        )

    # ========================================================
    # NSE/BSE CIRCULAR
    # ========================================================

    else:

        summary = (
            generate_exchange_circular_summary(
                text=text
            )
        )

        final_summary = (
            f"{title} dated "
            f"{issue_date}\n\n"
            f"{summary}"
        )

    row["Summary"] = final_summary

    row["EmbeddingText"] = text[:8000]

    return row

# ============================================================
# MAIN
# ============================================================

def main(
    excel_file: str
):

    df = pd.read_excel(
        excel_file
    )

    required = [
        "Verticals",
        "SubCategory",
        "Path"
    ]

    for col in required:

        if col not in df.columns:

            raise ValueError(
                f"Missing required column: "
                f"{col}"
            )

    logging.info(
        f"Total rows in input: "
        f"{len(df)}"
    )

    sebi_mask = (
        df["Verticals"]
        .str.strip()
        .str.lower()
        .isin(
            {
                "listed companies",
                "sebi",
                "aif"
            }
        )
    )

    df_sebi = (
        df[sebi_mask]
        .copy()
    )

    logging.info(
        f"SEBI rows to process: "
        f"{len(df_sebi)}"
    )

    start = time.time()

    processed_count = 0

    for idx, row in (
        df_sebi.iterrows()
    ):

        logging.info(
            f"[{idx+1}] Processing: "
            f"{row.get('Title', '')[:80]}"
        )

        processed = (
            process_newsletter_row(
                row
            )
        )

        if processed is None:

            continue

        update_excel(
            processed
        )

        processed_count += 1

    logging.info(
        f"Completed processing "
        f"{processed_count} "
        f"newsletter items in "
        f"{time.time() - start:.2f}s"
    )

    # ========================================================
    # MINIO UPLOAD
    # ========================================================

    try:

        minio = MinIOClient()

        month_folder_name = (
            MONTH_FOLDER.name
        )

        minio_prefix = (
            f"monthly_outputs/"
            f"{month_folder_name}/"
        )

        minio.delete_prefix(
            minio_prefix
        )

        for excel_name in CREATED_EXCELS:

            local_excel = (
                MONTH_FOLDER /
                excel_name
            )

            object_path = (
                f"{minio_prefix}"
                f"{excel_name}"
            )

            minio.upload_file(
                local_path=str(
                    local_excel
                ),
                object_path=object_path
            )

        logging.info(
            "Uploaded newsletter "
            "files to MinIO"
        )

    except Exception as e:

        logging.error(
            f"MinIO upload failed: "
            f"{e}"
        )

# ============================================================
# ENTRY
# ============================================================

if __name__ == "__main__":

    excel = (
        DATA_DIR /
        "Searching_agent_output.xlsx"
    )

    if not excel.exists():

        raise FileNotFoundError(
            "Searching_agent_output.xlsx "
            "not found"
        )

    main(str(excel))